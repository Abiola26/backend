"""
Utility functions for data processing, reporting, and file handling
"""
import pandas as pd
from io import BytesIO
import datetime
from typing import List
import secrets
import string

def generate_account_id(role: str = "user") -> str:
    """
    Generate a random account ID with letters, numbers, and special chars
    Prefix depends on role: ADM for admin, USE for user
    """
    alphabet = string.ascii_uppercase + string.digits + "#@&"
    prefix = "ADM-" if role == "admin" else "USE-"
    return prefix + "".join(secrets.choice(alphabet) for _ in range(8))


from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from app.schemas import FleetRecordOut, FleetSummary, DailySubtotal, DashboardStats, AnalyticsResponse
from app.models import FleetRecord, SystemSetting
from app.database import SessionLocal

def get_system_config() -> dict:
    """Helper to fetch system settings as a dict"""
    db = SessionLocal()
    try:
        settings = db.query(SystemSetting).all()
        return {s.key: s.value for s in settings}
    finally:
        db.close()


def calculate_remittance(revenue: float, fleet_code: str, config: dict = None) -> float:
    """
    Calculate remittance based on fleet code prefix.
    Supports dynamic config or defaults to hardcoded rules.
    """
    fleet_code = str(fleet_code).strip()
    
    # Check for dynamic config first
    if config:
        # Example config key: 'REMITTANCE_1' for 1xxx series
        prefix = fleet_code[0] if fleet_code else ''
        rate_key = f"REMITTANCE_{prefix}"
        if rate_key in config:
            try:
                rate = float(config[rate_key]) / 100
                return revenue * rate
            except:
                pass

    # Hardcoded Defaults
    if fleet_code.startswith("1"):
        return revenue * 0.84
    elif fleet_code.startswith("2"):
        return revenue * 0.875
    return revenue


class DataProcessor:
    """Handles data transformation and aggregation"""
    
    @staticmethod
    def detect_anomalies(df: pd.DataFrame) -> List[Anomaly]:
        """Detect unusual patterns in fleet data"""
        from ..schemas import Anomaly
        anomalies = []
        
        if df.empty:
            return []
            
        # 1. Statistical Outliers per Fleet (Z-score approach)
        for fleet in df['fleet'].unique():
            fleet_df = df[df['fleet'] == fleet]
            if len(fleet_df) < 5:
                continue
                
            mean = fleet_df['amount'].mean()
            std = fleet_df['amount'].std()
            
            if std == 0:
                continue
                
            for _, row in fleet_df.iterrows():
                z_score = abs(row['amount'] - mean) / std
                if z_score > 3:
                    anomalies.append(Anomaly(
                        date=row['date'],
                        fleet=row['fleet'],
                        amount=row['amount'],
                        reason=f"Significant deviation (Z-score: {z_score:.2f})",
                        severity="high"
                    ))
                elif z_score > 2:
                    anomalies.append(Anomaly(
                        date=row['date'],
                        fleet=row['fleet'],
                        amount=row['amount'],
                        reason="Unusual amount for this fleet",
                        severity="medium"
                    ))
                    
        return anomalies

    @staticmethod
    def process_analytics(records: List[FleetRecord]) -> AnalyticsResponse:
        """Process raw records into full analytics response"""
        from ..schemas import AnalyticsResponse, DashboardStats, FleetSummary, DailySubtotal, FleetRecordOut
        
        if not records:
            return AnalyticsResponse(
                records=[],
                fleet_summaries=[],
                daily_subtotals=[],
                dashboard_stats=DashboardStats(
                    total_revenue=0,
                    total_records=0,
                    top_performing_fleet="N/A",
                    average_trip_revenue=0
                ),
                anomalies=[]
            )

        # Convert to DataFrame for easy processing
        data = []
        for r in records:
            fleet_name = r.fleet.strip().upper()
            if fleet_name == "2010M":
                fleet_name = "2010"
                
            data.append({
                "id": r.id,
                "date": r.date,
                "fleet": fleet_name,
                "amount": r.amount
            })
        
        df = pd.DataFrame(data)
        df['date'] = pd.to_datetime(df['date']).dt.date
        
        # Fetch system config for dynamic calculations
        config = get_system_config()
        
        # 1. Fleet Summaries
        fleet_grp = df.groupby("fleet")["amount"].agg(["sum", "count"]).reset_index()
        summaries = []
        for _, row in fleet_grp.iterrows():
            fleet_code = str(row['fleet']).strip()
            revenue = row['sum']
            remittance = calculate_remittance(revenue, fleet_code, config)
                
            summaries.append(FleetSummary(
                fleet=fleet_code, 
                total_amount=revenue, 
                record_count=row['count'],
                remittance=remittance
            ))
        
        # 2. Daily Subtotals
        daily_grp = df.groupby(["date", "fleet"])["amount"].agg(["sum", "count"]).reset_index()
        daily_grp = daily_grp.sort_values(by=["date", "fleet"])
        
        subtotals = [
            DailySubtotal(date=row['date'], fleet=row['fleet'], daily_total=row['sum'], pax=row['count'])
            for _, row in daily_grp.iterrows()
        ]
        
        # 3. Dashboard Stats with Trend
        total_rev = df["amount"].sum()
        total_count = len(df)
        avg_rev = total_rev / total_count if total_count > 0 else 0
        
        top_fleet = "N/A"
        if not fleet_grp.empty:
            top_fleet = fleet_grp.loc[fleet_grp["sum"].idxmax(), "fleet"]
            
        # Calculate Trend (Last 7 days vs Previous 7 days)
        max_date = df['date'].max()
        last_7_days = df[df['date'] >= (max_date - pd.Timedelta(days=7))]
        prev_7_days = df[(df['date'] < (max_date - pd.Timedelta(days=7))) & (df['date'] >= (max_date - pd.Timedelta(days=14)))]
        
        rev_last = last_7_days['amount'].sum()
        rev_prev = prev_7_days['amount'].sum()
        
        trend_percent = 0.0
        if rev_prev > 0:
            trend_percent = ((rev_last - rev_prev) / rev_prev) * 100
            
        stats = DashboardStats(
            total_revenue=total_rev,
            total_records=total_count,
            top_performing_fleet=top_fleet,
            average_trip_revenue=avg_rev,
            revenue_trend_percent=trend_percent
        )
        
        # 4. Anomalies
        anomalies = DataProcessor.detect_anomalies(df)
        
        # 5. Records
        record_objs = [
            FleetRecordOut(id=d['id'], date=d['date'], fleet=d['fleet'], amount=d['amount'])
            for d in data
        ]
        
        return AnalyticsResponse(
            records=record_objs,
            fleet_summaries=summaries,
            daily_subtotals=subtotals,
            dashboard_stats=stats,
            anomalies=anomalies
        )


class ReportGenerator:
    """Generates Excel and PDF reports"""

    @staticmethod
    def generate_excel(analytics: AnalyticsResponse) -> BytesIO:
        """Create a multi-sheet Excel report with custom styling"""
        
        # Convert Pydantic models back to DataFrame for processing
        if not analytics.records:
             # Return empty excel if no data
            output = BytesIO()
            pd.DataFrame().to_excel(output)
            output.seek(0)
            return output

        records_data = [r.model_dump() for r in analytics.records]
        df = pd.DataFrame(records_data)
        
        # Ensure correct types
        df['date'] = pd.to_datetime(df['date']).dt.date
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
        df['fleet'] = df['fleet'].astype(str)

        # ---------------- PREPARE DATA ----------------
        
        # 1. Daily Subtotals Logic
        grouped = df.groupby(["date", "fleet"]).agg(
            TotalAmount=("amount", "sum"),
            FleetCount=("fleet", "count")
        ).reset_index()
        
        # Explicit Sort
        grouped = grouped.sort_values(by=["date", "fleet"])

        formatted_rows = []
        for date_val, group in grouped.groupby("date"):
            sub_amt = group["TotalAmount"].sum()
            sub_cnt = group["FleetCount"].sum()

            for _, row in group.iterrows():
                formatted_rows.append({
                    "Date": row["date"],
                    "BUS CODE": row["fleet"],
                    "PAX": row["FleetCount"],
                    "REVENUE": row["TotalAmount"]
                })

            # Subtotal Row
            formatted_rows.append({
                "Date": date_val,
                "BUS CODE": "Subtotal",
                "PAX": sub_cnt,
                "REVENUE": sub_amt
            })
            
        subtotal_df = pd.DataFrame(formatted_rows)
        
        # 2. Bus Code Performance Logic (formerly Fleet Summary)
        bus_summary_df = df.groupby("fleet").agg(
            PAX=("fleet", "count"),
            REVENUE=("amount", "sum")
        ).reset_index().rename(columns={"fleet": "BUS CODE"})
        
        bus_summary_df = bus_summary_df.sort_values("BUS CODE")

        # Fetch config for excel gen
        config = get_system_config()

        # Calculate REMITTANCE and FUEL USED
        def calculate_metrics(row):
            code = str(row["BUS CODE"]).strip()
            revenue = row["REVENUE"]
            
            remittance = calculate_remittance(revenue, code, config)
                
            # Placeholder logic for Fuel Used since not in DB
            # Matches roughly the ratio in image for visual completeness
            fuel_used = revenue * 0.30 
            
            return pd.Series([remittance, fuel_used])

        bus_summary_df[["REMITTANCE", "FUEL USED"]] = bus_summary_df.apply(calculate_metrics, axis=1)

        # Add Grand Total
        bus_summary_df.loc[len(bus_summary_df)] = [
            "Grand Total",
            bus_summary_df["PAX"].sum(),
            bus_summary_df["REVENUE"].sum(),
            bus_summary_df["REMITTANCE"].sum(),
            bus_summary_df["FUEL USED"].sum()
        ]

        # 3. Dashboard Stats Logic
        stats_df = pd.DataFrame({
            "Metric": ["Total Revenue", "Total Records", "Top Fleet", "Avg Revenue"],
            "Value": [
                analytics.dashboard_stats.total_revenue,
                analytics.dashboard_stats.total_records,
                analytics.dashboard_stats.top_performing_fleet,
                analytics.dashboard_stats.average_trip_revenue
            ]
        })

        # ---------------- WRITE TO EXCEL ----------------
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Bus Performance (Primary Report)
            bus_summary_df.to_excel(writer, sheet_name="Bus Performance", index=False)
            
            # Sheet 2: Dashboard Stats
            stats_df.to_excel(writer, sheet_name="Dashboard", index=False)
            
            # Sheet 3: Daily Subtotals
            subtotal_df.to_excel(writer, sheet_name="Daily Subtotals", index=False)
            
            # Sheet 4: Raw Data
            df.to_excel(writer, sheet_name="Raw Data", index=False)

        # ---------------- APPLY STYLING ----------------
        output.seek(0)
        wb = load_workbook(output)
        
        # Define Styles ONCE to optimize performance (reuse objects)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
        header_font = Font(bold=True, color="FFFFFF")
        
        subtotal_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid") # Darker Grey for Total
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
        blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        def style_worksheet(ws, sheet_name):
            # Header Styling
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
            # Row Styling
            # Optimize: check sheet name once
            is_bus_perf = (sheet_name == "Bus Performance")
            is_daily = (sheet_name == "Daily Subtotals")
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                 # Apply Borders to ALL cells
                 for cell in row:
                     cell.border = thin_border

                 # Conditional Formatting
                 bus_code = None
                 if is_bus_perf:
                     bus_code = str(row[0].value) 
                     
                     if bus_code == "Grand Total" or bus_code == "Subtotal":
                         for cell in row: 
                             cell.font = bold_font
                             cell.fill = subtotal_fill
                     elif bus_code.startswith("1"):
                         for cell in row: cell.fill = yellow_fill
                     elif bus_code.startswith("2"):
                         for cell in row: cell.fill = blue_fill
                         
                 elif is_daily:
                     bus_code = str(row[1].value) 
                     if bus_code == "Subtotal":
                         for cell in row:
                             cell.font = bold_font
                             cell.fill = subtotal_fill

                 # Currency & Number Format
                 for cell in row:
                    if isinstance(cell.value, (int, float)):
                         # We can check header by index instead of lookup to save time
                         # But for now, simple optimization: cache header row? 
                         # ws[1][cell.col_idx - 1] ... 
                         # Let's keep heuristic but optimize string ops
                         header_val = ws.cell(row=1, column=cell.column).value
                         header_str = str(header_val).upper()
                         
                         if "REVENUE" in header_str or "REMITTANCE" in header_str or "AMOUNT" in header_str or "FUEL" in header_str:
                             cell.number_format = '#,##0.00'
                         elif "PAX" in header_str or "COUNT" in header_str:
                             cell.number_format = '#,##0'

            # Auto-adjust columns
            for column_cells in ws.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Apply to all sheets
        for sheet_name in wb.sheetnames:
            style_worksheet(wb[sheet_name], sheet_name)
            
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        return final_output

    @staticmethod
    def generate_pdf(analytics: AnalyticsResponse) -> BytesIO:
        """Create a PDF report"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph("Fleet Reporting System - Analytical Report", styles['Title']))
        elements.append(Spacer(1, 12))
        
        # Summary Section
        elements.append(Paragraph("Executive Summary", styles['Heading2']))
        stats = analytics.dashboard_stats
        summary_data = [
            ["Metric", "Value"],
            ["Total Revenue", f"{stats.total_revenue:,.2f}"],
            ["Total Records", f"{stats.total_records}"],
            ["Top Fleet", stats.top_performing_fleet],
            ["Avg Revenue/Trip", f"{stats.average_trip_revenue:,.2f}"]
        ]
        
        t = Table(summary_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        # Fleet Breakdown
        elements.append(Paragraph("Fleet Performance Breakdown", styles['Heading2']))
        if analytics.fleet_summaries:
            data = [["Fleet", "Total Revenue", "Count"]]
            for s in analytics.fleet_summaries:
                data.append([s.fleet, f"{s.total_amount:,.2f}", str(s.record_count)])
                
            t2 = Table(data)
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ]))
            elements.append(t2)
        else:
            elements.append(Paragraph("No data available", styles['Normal']))

        doc.build(elements)
        output.seek(0)
        return output
